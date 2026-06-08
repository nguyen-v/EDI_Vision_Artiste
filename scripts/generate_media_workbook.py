#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate MediaMap_workbook.xlsx for Vision Artiste quiz (English UI).

Sheets:
  Guide       - how to use, limits, tie-break
  Questions   - quiz mapping (editable screen + score grid); drives CodeGen
  SD WIZ1     - files for portrait display (1080x1920)
  SD WIZ2     - files for landscape display (1920x1080)
  CodeGen     - paste block for QuestionnaireConfig.h

Usage:
    python scripts/generate_media_workbook.py
    python scripts/generate_media_workbook.py --emit-cpp
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

NUM_LANGUAGES = 4
FILES_PER_LANGUAGE = 32
MEDEWIZ_MAX_FILES = 200
NUM_CHOICES = 4
NUM_CATEGORIES = 4
# Initial quiz steps baked into a new workbook (edit or add rows on Questions sheet).
DEFAULT_NUM_QUESTIONS = 13
# Extra blank rows with formulas ready (fill the next row to add a step).
MAX_QUESTION_SLOTS = 24
LANG_CODES = ("EN", "FR", "DE", "IT")
CAT_NAMES = ("Emotions", "Realiste", "Matiere", "Conteur")
CAT_CPP = ("CAT_EMOTIONS", "CAT_REALISTE", "CAT_MATIERE", "CAT_CONTEUR")

# Per-answer points per profile (0–10 on the Questions score grid).
SCORE_WEIGHT_MIN = 0
SCORE_WEIGHT_MAX = 10
SCORE_WEIGHT_PRIMARY = 10
SCORE_WEIGHT_OTHER = 0

QUESTION_FILE_FIRST_OFFSET = 1
# Artwork uses the same SD index as question text offset (separate cards, no collision).
QUESTION_IMAGE_OFFSET_DELTA = 0

QUESTION_SCREEN_ORDER = "lppllllpllplp"

PORTRAIT_ASPECT = "1080x1920"
LANDSCAPE_ASPECT = "1920x1080"

QUESTIONS_FIRST_ROW = 4
QUESTIONS_LAST_ROW = QUESTIONS_FIRST_ROW + MAX_QUESTION_SLOTS - 1

# Questions columns (1-based)
COL_QNUM = 1
COL_ARTWORK = 2
COL_QUESTION_ON = 3
COL_ARTWORK_ON = 4
COL_TEXT_OFFSET = 5
COL_IMAGE_INDEX = 6
COL_TEXT_PLAYER = 7
COL_ART_PLAYER = 8
COL_TEXT_ASPECT = 9
COL_ART_ASPECT = 10
COL_TEXT_FILE_FR = 11
COL_ART_FILE = 12
# B1..B4 × Emotions, Realiste, Matiere, Conteur (editable 0–10).
COL_WEIGHTS = 13
COL_WEIGHTS_LAST = COL_WEIGHTS + NUM_CHOICES * NUM_CATEGORIES - 1

HIDDEN_LINE_COL = "AC"
COL_HIDDEN_LINE = column_index_from_string(HIDDEN_LINE_COL)

CODEGEN_BEGIN_ROW = 4
CODEGEN_NUM_QUESTIONS_ROW = 5
CODEGEN_ARRAY_OPEN_ROW = 6
CODEGEN_FIRST_Q_ROW = 7
CODEGEN_LAST_Q_ROW = CODEGEN_FIRST_Q_ROW + MAX_QUESTION_SLOTS - 1
CODEGEN_CLOSE_ROW = CODEGEN_LAST_Q_ROW + 1
CODEGEN_END_MARKER_ROW = CODEGEN_CLOSE_ROW + 1
COPY_BLOCK_FIRST_ROW = 4
COPY_BLOCK_LAST_ROW = CODEGEN_END_MARKER_ROW

WIZ1_FIRST_ROW = 4
WIZ2_FIRST_ROW = 4

TITLE_FILL = PatternFill("solid", fgColor="2F5496")
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
INSTRUCTION_FILL = PatternFill("solid", fgColor="F8F9FA")
INSTRUCTION_FONT = Font(size=10, color="404040")
EDIT_FILL = PatternFill("solid", fgColor="E2EFDA")
EDIT_HEADER_FILL = PatternFill("solid", fgColor="C6E0B4")
INPUT_FILL = PatternFill("solid", fgColor="F2F2F2")
INPUT_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
CALC_FILL = PatternFill("solid", fgColor="DDEBF7")
FILE_HEADER_FILL = PatternFill("solid", fgColor="9BC2E6")
COPY_FILL = PatternFill("solid", fgColor="FFF2CC")
# B1–B4 score columns (header, editable cell, calc/read-only in block).
BUTTON_HEADER_FILLS = (
    PatternFill("solid", fgColor="9DC3E6"),  # B1 blue
    PatternFill("solid", fgColor="A9D08E"),  # B2 green
    PatternFill("solid", fgColor="F4B084"),  # B3 orange
    PatternFill("solid", fgColor="C9B8D8"),  # B4 purple
)
BUTTON_DATA_FILLS = (
    PatternFill("solid", fgColor="DDEBF7"),
    PatternFill("solid", fgColor="E2EFDA"),
    PatternFill("solid", fgColor="FCE4D6"),
    PatternFill("solid", fgColor="E4DFEC"),
)
BUTTON_CALC_FILLS = (
    PatternFill("solid", fgColor="EDF4FC"),
    PatternFill("solid", fgColor="F4FAF0"),
    PatternFill("solid", fgColor="FEF3EC"),
    PatternFill("solid", fgColor="F5F0FA"),
)
SECTION_FONT = Font(bold=True, size=11, color="2F5496")
HEADER_FONT = Font(bold=True, color="1F3864")
DATA_FONT = Font(size=11)
DATA_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
UNLOCKED = Protection(locked=False)
LOCKED = Protection(locked=True)

SCREEN_LIST = "Portrait,Landscape"


def screen_from_order_char(ch: str) -> str:
    return "Landscape" if ch.lower() == "l" else "Portrait"


_SCREEN = [screen_from_order_char(c) for c in QUESTION_SCREEN_ORDER]
assert len(_SCREEN) == DEFAULT_NUM_QUESTIONS

# (default B1..B4 primaries, question_on_screen, artwork_title) — primaries seed score grid only.
DEFAULT_QUESTIONS: Tuple[Tuple[Tuple[int, ...], str, str], ...] = (
    ((0, 1, 2, 3), _SCREEN[0], "Caravaggio - Supper at Emmaus"),
    ((3, 1, 2, 0), _SCREEN[1], "Monet - Rouen Cathedrals"),
    ((0, 1, 2, 3), _SCREEN[2], "Magritte - Empire of Light"),
    ((0, 1, 2, 3), _SCREEN[3], "Delaunay / Kirchner - night landscape"),
    ((0, 1, 2, 3), _SCREEN[4], "Van Gogh - Starry Night"),
    ((0, 3, 2, 1), _SCREEN[5], "Hopper - Nighthawks (Q6A)"),
    ((1, 2, 3, 0), _SCREEN[6], "Hopper - Night Windows (Q6B)"),
    ((2, 3, 1, 0), _SCREEN[7], "Photo - Sudek / Brassai (Q7)"),
    ((0, 3, 1, 2), _SCREEN[8], "Fuseli - The Nightmare"),
    ((2, 3, 1, 0), _SCREEN[9], "Soulages - Outrenoir"),
    ((0, 3, 1, 2), _SCREEN[10], "Vallotton - The Bibliophile"),
    ((1, 0, 2, 3), _SCREEN[11], "Grandville - Cast Shadows"),
    ((0, 3, 1, 2), _SCREEN[12], "Your work - final 4 motifs Q12"),
)

WIZ1_FIXED_ROWS = (
    ("Idle intro text", "-", 0, "Title + intro (all languages)", PORTRAIT_ASPECT),
    ("Profile Emotions", "-", 17, "Result text - poetic / emotions", PORTRAIT_ASPECT),
    ("Profile Realiste", "-", 18, "Result text - precision", PORTRAIT_ASPECT),
    ("Profile Matiere", "-", 19, "Result text - matter", PORTRAIT_ASPECT),
    ("Profile Conteur", "-", 20, "Result text - storyteller", PORTRAIT_ASPECT),
)

WIZ2_FIXED_ROWS = (
    ("Idle artwork", "-", 0, "Intro visual landscape (all languages)", LANDSCAPE_ASPECT),
    ("Profile art Emotions", "-", 17, "Result visual - emotions", LANDSCAPE_ASPECT),
    ("Profile art Realiste", "-", 18, "Result visual - precision", LANDSCAPE_ASPECT),
    ("Profile art Matiere", "-", 19, "Result visual - matter", LANDSCAPE_ASPECT),
    ("Profile art Conteur", "-", 20, "Result visual - storyteller", LANDSCAPE_ASPECT),
)


def role_has_language_files(role: str) -> bool:
    if role.startswith("Question text") or role.startswith("Idle intro"):
        return True
    return role.startswith("Profile") and not role.startswith("Profile art")


def placement(question_on: str) -> Tuple[str, str, str, str]:
    """Return text_wiz, art_wiz, text_aspect, art_aspect."""
    if question_on == "Portrait":
        return "WIZ1", "WIZ2", PORTRAIT_ASPECT, LANDSCAPE_ASPECT
    return "WIZ2", "WIZ1", LANDSCAPE_ASPECT, PORTRAIT_ASPECT


def read_question_entries_from_ws(ws: Worksheet) -> List[Tuple[int, str, str, int, int]]:
    """Filled Questions rows: (step, question_on, artwork, text_offset, image_index)."""
    entries: List[Tuple[int, str, str, int, int]] = []
    for row in iter_question_rows_from_sheet(ws):
        artwork = str(ws.cell(row=row, column=COL_ARTWORK).value).strip()
        step_val = ws.cell(row=row, column=COL_QNUM).value
        try:
            step = int(step_val)
        except (TypeError, ValueError):
            step = row - QUESTIONS_FIRST_ROW + 1
        question_on = str(ws.cell(row=row, column=COL_QUESTION_ON).value or "Portrait").strip()
        text_off = ws.cell(row=row, column=COL_TEXT_OFFSET).value
        img_idx = ws.cell(row=row, column=COL_IMAGE_INDEX).value
        try:
            text_offset = int(text_off)
        except (TypeError, ValueError):
            text_offset = step
        try:
            image_index = int(img_idx)
        except (TypeError, ValueError):
            image_index = text_offset
        entries.append((step, question_on, artwork, text_offset, image_index))
    return entries


def build_sd_slot_rows_from_entries(
    entries: Sequence[Tuple[int, str, str, int, int]],
) -> Tuple[List[Tuple], List[Tuple]]:
    """Build SD WIZ1/WIZ2 row tuples (sorted by file index) from Questions entries."""
    wiz1: List[Tuple] = list(WIZ1_FIXED_ROWS)
    wiz2: List[Tuple] = list(WIZ2_FIXED_ROWS)

    for step, question_on, artwork, text_offset, image_index in entries:
        qref = f"Q{step}"
        text_wiz, art_wiz, text_asp, art_asp = placement(question_on)
        q_screen = "portrait" if question_on == "Portrait" else "landscape"
        art_screen = "portrait" if art_wiz == "WIZ1" else "landscape"

        if text_wiz == "WIZ1":
            wiz1.append((
                f"Question text {qref}",
                qref,
                text_offset,
                f"{artwork} (question on {q_screen} screen)",
                text_asp,
            ))
        else:
            wiz2.append((
                f"Question text {qref}",
                qref,
                text_offset,
                f"{artwork} (question on {q_screen} screen)",
                text_asp,
            ))

        if art_wiz == "WIZ1":
            wiz1.append((
                f"Artwork {qref}",
                qref,
                image_index,
                f"{artwork} (art on {art_screen} screen)",
                art_asp,
            ))
        else:
            wiz2.append((
                f"Artwork {qref}",
                qref,
                image_index,
                f"{artwork} (art on {art_screen} screen)",
                art_asp,
            ))

    wiz1.sort(key=lambda r: (r[2] if isinstance(r[2], int) else 999))
    wiz2.sort(key=lambda r: (r[2] if isinstance(r[2], int) else 999))
    return wiz1, wiz2


def build_sd_slot_rows() -> Tuple[List[Tuple], List[Tuple]]:
    """Default SD rows from DEFAULT_QUESTIONS (used when no workbook is loaded)."""
    entries = [
        (i + 1, question_on, artwork, i + 1, i + 1 + QUESTION_IMAGE_OFFSET_DELTA)
        for i, (_cats, question_on, artwork) in enumerate(DEFAULT_QUESTIONS)
    ]
    return build_sd_slot_rows_from_entries(entries)


def sd_rows_from_workbook(path: Path) -> Tuple[List[Tuple], List[Tuple]]:
    """SD row lists derived from the Questions sheet in a saved workbook."""
    wb = load_workbook(path, data_only=True)
    if "Questions" not in wb.sheetnames:
        raise ValueError(f"Questions sheet missing in {path}")
    return build_sd_slot_rows_from_entries(read_question_entries_from_ws(wb["Questions"]))


WIZ1_SD_ROWS, WIZ2_SD_ROWS = build_sd_slot_rows()
SD_DYNAMIC_ROWS = MAX_QUESTION_SLOTS * 2
WIZ1_LAST_ROW = WIZ1_FIRST_ROW + len(WIZ1_FIXED_ROWS) + SD_DYNAMIC_ROWS - 1
WIZ2_LAST_ROW = WIZ2_FIRST_ROW + len(WIZ2_FIXED_ROWS) + SD_DYNAMIC_ROWS - 1


def set_col_widths(ws: Worksheet, widths: Sequence[float]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def apply_cell_style(cell, *, fill=None, font=None, alignment=None, border=None, locked=True):
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    cell.protection = LOCKED if locked else UNLOCKED


def enable_sheet_protection(ws: Worksheet) -> None:
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False


def button_choice_index(col: int) -> int | None:
    """0..3 for B1–B4 score block columns, else None."""
    if COL_WEIGHTS <= col <= COL_WEIGHTS_LAST:
        return (col - COL_WEIGHTS) // NUM_CATEGORIES
    return None


def questions_sheet_fill(
    col: int,
    *,
    editable: bool,
    calc: bool,
    header: bool,
) -> PatternFill:
    choice = button_choice_index(col)
    if choice is not None:
        if header:
            return BUTTON_HEADER_FILLS[choice]
        if editable:
            return BUTTON_DATA_FILLS[choice]
        return BUTTON_CALC_FILLS[choice]
    if col in (COL_TEXT_FILE_FR, COL_ART_FILE):
        return FILE_HEADER_FILL if header else CALC_FILL
    if calc:
        return CALC_FILL
    if editable:
        return EDIT_FILL
    return INPUT_FILL


def style_merged_title(ws, cell_ref, merge_range):
    ws.merge_cells(merge_range)
    cell = ws[cell_ref]
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_instruction_row(ws, cell_ref, merge_range):
    ws.merge_cells(merge_range)
    cell = ws[cell_ref]
    cell.font = INSTRUCTION_FONT
    cell.fill = INSTRUCTION_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_list_validation(ws: Worksheet, cell_range: str, options: str) -> None:
    dv = DataValidation(
        type="list",
        formula1=f'"{options}"',
        allow_blank=False,
        showDropDown=False,  # False = show arrow in Excel / Calc
    )
    dv.error = "Choose a value from the list."
    dv.errorTitle = "Invalid value"
    dv.add(cell_range)
    ws.add_data_validation(dv)


def lang_filename_formula(row: int, lang_idx: int, index_col: str = "C") -> str:
    ref = f"{index_col}{row}"
    return (
        f'=IF({ref}="","",'
        f'TEXT({FILES_PER_LANGUAGE}*{lang_idx}+{ref},"000")&".mp4")'
    )


def file_index_formula(row: int, index_col: str = "C") -> str:
    ref = f"{index_col}{row}"
    return f'=IF({ref}="","",TEXT({ref},"000")&".mp4")'


def default_weights_for_primary(primary: int) -> Tuple[int, int, int, int]:
    """Default template for one button: primary profile = 10, others = 0."""
    weights = [SCORE_WEIGHT_OTHER] * NUM_CATEGORIES
    if 0 <= primary < NUM_CATEGORIES:
        weights[primary] = SCORE_WEIGHT_PRIMARY
    return tuple(weights)


def default_weights_for_question(
    cats: Sequence[int],
) -> Tuple[Tuple[int, int, int, int], ...]:
    return tuple(default_weights_for_primary(cats[b]) for b in range(NUM_CHOICES))


def zero_weights_for_choice() -> Tuple[int, int, int, int]:
    return (0, 0, 0, 0)


def cpp_question_line(
    offset: int,
    screen: str,
    weights: Sequence[Sequence[int]],
) -> str:
    rows = ", ".join(
        "{ " + ", ".join(str(weights[ch][c]) for c in range(NUM_CATEGORIES)) + " }"
        for ch in range(NUM_CHOICES)
    )
    return f"  {{ {offset}, {screen}, {{ {rows} }} }},"


def cpp_question_line_formula(row: int) -> str:
    """Live C++ QUESTIONS[] entry from offset, screen, and score grid (column AC)."""
    art = f"{get_column_letter(COL_ARTWORK)}{row}"
    off = f"{get_column_letter(COL_TEXT_OFFSET)}{row}"
    q_on = f"{get_column_letter(COL_QUESTION_ON)}{row}"
    screen = f'IF({q_on}="Portrait","Portrait","Landscape")'
    button_rows: List[str] = []
    for choice in range(NUM_CHOICES):
        cells = [
            f"{get_column_letter(COL_WEIGHTS + choice * NUM_CATEGORIES + cat)}{row}"
            for cat in range(NUM_CATEGORIES)
        ]
        button_rows.append('" { "&' + '&", "&'.join(cells) + '&" }"')
    weights = '"{ "&' + '&", "&'.join(button_rows) + '&" }"'
    return (
        f'=IF({art}="","",'
        f'"  {{ "&{off}&", "&{screen}&", "&{weights}&" }} }},")'
    )


def step_formula(row: int) -> str:
    art = f"{get_column_letter(COL_ARTWORK)}{row}"
    return f'=IF({art}="","",ROW()-{QUESTIONS_FIRST_ROW - 1})'


def text_offset_formula(row: int) -> str:
    """SD text file index (= step number); hidden from media team."""
    step_col = get_column_letter(COL_QNUM)
    art = f"{get_column_letter(COL_ARTWORK)}{row}"
    return f'=IF({art}="","",{step_col}{row})'


def count_questions_formula() -> str:
    art_col = get_column_letter(COL_ARTWORK)
    return (
        f"COUNTA(Questions!${art_col}${QUESTIONS_FIRST_ROW}:"
        f"${art_col}${QUESTIONS_LAST_ROW})"
    )


def num_questions_cpp_formula() -> str:
    return f'="static const uint8_t NUM_QUESTIONS = "&{count_questions_formula()}&";"'


def fill_question_row(
    ws: Worksheet,
    row: int,
    *,
    default_cats: Sequence[int] | None = None,
    question_on: str | None = None,
    artwork: str | None = None,
) -> None:
    """Write one Questions row (formulas + optional default values)."""
    q_col = get_column_letter(COL_QUESTION_ON)
    off_col = get_column_letter(COL_TEXT_OFFSET)
    img_col = get_column_letter(COL_IMAGE_INDEX)

    ws.cell(row=row, column=COL_QNUM, value=step_formula(row))
    if artwork is not None:
        ws.cell(row=row, column=COL_ARTWORK, value=artwork)
    if question_on is not None:
        ws.cell(row=row, column=COL_QUESTION_ON, value=question_on)
    ws.cell(
        row=row, column=COL_ARTWORK_ON,
        value=f'=IF({q_col}{row}="Portrait","Landscape","Portrait")',
    )
    ws.cell(row=row, column=COL_TEXT_OFFSET, value=text_offset_formula(row))
    ws.cell(
        row=row, column=COL_IMAGE_INDEX,
        value=f"={off_col}{row}+{QUESTION_IMAGE_OFFSET_DELTA}",
    )
    if default_cats is not None:
        weights = default_weights_for_question(default_cats)
        for choice in range(NUM_CHOICES):
            w = weights[choice]
            for cat in range(NUM_CATEGORIES):
                ws.cell(
                    row=row,
                    column=COL_WEIGHTS + choice * NUM_CATEGORIES + cat,
                    value=w[cat],
                )
    ws.cell(
        row=row, column=COL_TEXT_PLAYER,
        value=f'=IF({q_col}{row}="Portrait","WIZ1","WIZ2")',
    )
    ws.cell(
        row=row, column=COL_ART_PLAYER,
        value=f'=IF({q_col}{row}="Portrait","WIZ2","WIZ1")',
    )
    ws.cell(
        row=row, column=COL_TEXT_ASPECT,
        value=f'=IF({q_col}{row}="Portrait","{PORTRAIT_ASPECT}","{LANDSCAPE_ASPECT}")',
    )
    ws.cell(
        row=row, column=COL_ART_ASPECT,
        value=f'=IF({q_col}{row}="Portrait","{LANDSCAPE_ASPECT}","{PORTRAIT_ASPECT}")',
    )
    ws.cell(
        row=row, column=COL_TEXT_FILE_FR,
        value=f'=IF({get_column_letter(COL_ARTWORK)}{row}="","",TEXT(32+{off_col}{row},"000")&".mp4")',
    )
    ws.cell(
        row=row, column=COL_ART_FILE,
        value=f'=IF({get_column_letter(COL_ARTWORK)}{row}="","",TEXT({img_col}{row},"000")&".mp4")',
    )
    ws.cell(row=row, column=COL_HIDDEN_LINE, value=cpp_question_line_formula(row))


def codegen_question_ref(row: int) -> str:
    return f"='Questions'!${HIDDEN_LINE_COL}{row}"


def build_guide_sheet(ws: Worksheet) -> None:
    ws.title = "Guide"
    ws.sheet_properties.tabColor = "2F5496"
    style_merged_title(ws, "A1", "A1:D1")
    ws["A1"] = "Vision Artiste - MedeaWiz media map"
    lines = [
        "Two MedeaWiz Sprite players:",
        "  WIZ1 - portrait display (1080 x 1920)",
        "  WIZ2 - landscape display (1920 x 1080)",
        "",
        "For each quiz step (Questions sheet):",
        "  - Pick where the QUESTION TEXT plays: Portrait or Landscape (dropdown).",
        "  - The ARTWORK always plays on the other screen (complementary).",
        "  - Export video with the matching aspect ratio on each SD card.",
        "",
        "SD file names: 000.mp4, 001.mp4, ... (3 digits + .mp4).",
        "Index = file number sent to the Sprite.",
        "",
        "Flow: intro (text + art) -> quiz steps -> profile on WIZ1.",
        "Scoring: score grid (columns M–AB) — 0–10 points per profile per button (B1–B4).",
        "  Sum across steps; highest wins. Tie-break: Emotions, Conteur, Matiere, Realiste.",
        "",
        "Add a question: on Questions sheet, fill the next empty row (Artwork column).",
        "  Clear Artwork to remove a step (do not delete rows — CodeGen uses fixed row refs).",
        "  CodeGen column A updates when you edit scores, screen, or Artwork (recalc in Excel).",
        "  Or run: python scripts/generate_media_workbook.py --emit-cpp",
        "  SD WIZ1 / SD WIZ2 update automatically from Questions (clear Artwork to remove a step).",
        "",
        "Sheets: Questions | SD WIZ1 | SD WIZ2 | CodeGen",
        "",
        "If CodeGen shows Err:508 in LibreOffice, run: python scripts/generate_media_workbook.py --emit-cpp",
    ]
    first_limit_row = 3 + len(lines) + 1
    for i, line in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=line)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
    ws.cell(row=first_limit_row, column=1, value="Firmware limits").font = SECTION_FONT
    for j, (lab, val) in enumerate([
        ("Files per language (question text)", str(FILES_PER_LANGUAGE)),
        ("Languages", str(NUM_LANGUAGES)),
        ("Max files per SD card", str(MEDEWIZ_MAX_FILES)),
        ("Question rows in workbook", str(MAX_QUESTION_SLOTS)),
        ("Quiz steps (initial)", str(DEFAULT_NUM_QUESTIONS)),
    ]):
        row = first_limit_row + 1 + j
        ws.cell(row=row, column=1, value=lab)
        ws.cell(row=row, column=2, value=val)
    set_col_widths(ws, (58, 18, 12, 12))
    for row in range(3, first_limit_row):
        apply_cell_style(ws.cell(row=row, column=1), fill=INSTRUCTION_FILL, locked=True)
    enable_sheet_protection(ws)


def _questions_cell(row: int, col: int) -> str:
    """Excel reference to a Questions sheet cell (sheet name contains a space)."""
    return f"'Questions'!${get_column_letter(col)}${row}"


def _sd_text_active(qrow: int, wiz_player: str) -> str:
    art = _questions_cell(qrow, COL_ARTWORK)
    player = _questions_cell(qrow, COL_TEXT_PLAYER)
    return f'AND({art}<>"",{player}="{wiz_player}")'


def _sd_art_active(qrow: int, wiz_player: str) -> str:
    art = _questions_cell(qrow, COL_ARTWORK)
    player = _questions_cell(qrow, COL_ART_PLAYER)
    return f'AND({art}<>"",{player}="{wiz_player}")'


def _write_sd_static_row(
    ws: Worksheet,
    row: int,
    row_data: Tuple,
) -> None:
    role, qref, file_idx, desc, row_aspect = row_data
    ws.cell(row=row, column=1, value=role)
    ws.cell(row=row, column=2, value=qref)
    ws.cell(row=row, column=3, value=file_idx)
    ws.cell(row=row, column=4, value=desc)
    ws.cell(row=row, column=5, value=row_aspect)
    if role_has_language_files(role):
        for lang_idx in range(NUM_LANGUAGES):
            ws.cell(row=row, column=6 + lang_idx, value=lang_filename_formula(row, lang_idx))
    else:
        ws.cell(row=row, column=6, value=file_index_formula(row))


def _write_sd_question_text_row(ws: Worksheet, row: int, qrow: int, wiz_player: str) -> None:
    active = _sd_text_active(qrow, wiz_player)
    step = _questions_cell(qrow, COL_QNUM)
    art = _questions_cell(qrow, COL_ARTWORK)
    q_on = _questions_cell(qrow, COL_QUESTION_ON)
    off = _questions_cell(qrow, COL_TEXT_OFFSET)
    asp = _questions_cell(qrow, COL_TEXT_ASPECT)
    screen = f'IF({q_on}="Portrait","portrait","landscape")'
    ws.cell(row=row, column=1, value=f'=IF({active},"Question text Q"&{step},"")')
    ws.cell(row=row, column=2, value=f'=IF({active},"Q"&{step},"")')
    ws.cell(row=row, column=3, value=f"=IF({active},{off},\"\")")
    ws.cell(
        row=row, column=4,
        value=f'=IF({active},{art}&" (question on "&{screen}&" screen)","")',
    )
    ws.cell(row=row, column=5, value=f"=IF({active},{asp},\"\")")
    for lang_idx in range(NUM_LANGUAGES):
        ws.cell(
            row=row,
            column=6 + lang_idx,
            value=(
                f'=IF({active},TEXT({FILES_PER_LANGUAGE}*{lang_idx}+'
                f'{off},"000")&".mp4","")'
            ),
        )


def _write_sd_artwork_row(ws: Worksheet, row: int, qrow: int, wiz_player: str) -> None:
    active = _sd_art_active(qrow, wiz_player)
    step = _questions_cell(qrow, COL_QNUM)
    art = _questions_cell(qrow, COL_ARTWORK)
    img = _questions_cell(qrow, COL_IMAGE_INDEX)
    asp = _questions_cell(qrow, COL_ART_ASPECT)
    art_screen = "portrait" if wiz_player == "WIZ1" else "landscape"
    ws.cell(row=row, column=1, value=f'=IF({active},"Artwork Q"&{step},"")')
    ws.cell(row=row, column=2, value=f'=IF({active},"Q"&{step},"")')
    ws.cell(row=row, column=3, value=f"=IF({active},{img},\"\")")
    ws.cell(
        row=row, column=4,
        value=f'=IF({active},{art}&" (art on {art_screen} screen)","")',
    )
    ws.cell(row=row, column=5, value=f"=IF({active},{asp},\"\")")
    ws.cell(row=row, column=6, value=f'=IF({active},TEXT({img},"000")&".mp4","")')


def build_wiz_sd_sheet(
    ws: Worksheet,
    *,
    title: str,
    wiz_name: str,
    wiz_player: str,
    fixed_rows: Sequence[Tuple],
    tab_color: str,
    first_row: int,
    last_row: int,
) -> None:
    ws.title = title
    ws.sheet_properties.tabColor = tab_color
    style_merged_title(ws, "A1", "A1:I1")
    ws["A1"] = f"SD card - {wiz_name}"
    ws["A2"] = (
        "Copy each file to this SD card using the exact name shown. "
        "Encode each row at the Aspect shown (portrait 1080x1920 or landscape 1920x1080). "
        "Quiz rows are computed from the Questions sheet; clear Artwork there to drop a step."
    )
    style_instruction_row(ws, "A2", "A2:I2")
    headers = [
        "Role", "Quiz ref", "File index", "Content", "Aspect",
        "EN file", "FR file", "DE file", "IT file",
    ]
    for col, hdr in enumerate(headers, start=1):
        fill = EDIT_HEADER_FILL if col == 3 else (
            FILE_HEADER_FILL if col >= 6 else INPUT_HEADER_FILL
        )
        apply_cell_style(ws.cell(row=3, column=col, value=hdr), fill=fill, font=HEADER_FONT, locked=True)

    for i, row_data in enumerate(fixed_rows):
        _write_sd_static_row(ws, first_row + i, row_data)

    row = first_row + len(fixed_rows)
    for slot in range(MAX_QUESTION_SLOTS):
        qrow = QUESTIONS_FIRST_ROW + slot
        _write_sd_question_text_row(ws, row, qrow, wiz_player)
        row += 1
        _write_sd_artwork_row(ws, row, qrow, wiz_player)
        row += 1

    center = Alignment(horizontal="center", vertical="center")
    dynamic_start = first_row + len(fixed_rows)
    for row in range(first_row, last_row + 1):
        is_fixed = row < dynamic_start
        if is_fixed:
            role = ws.cell(row=row, column=1).value or ""
            has_langs = role_has_language_files(str(role))
        else:
            has_langs = (row - dynamic_start) % 2 == 0
        for col in range(1, 10):
            if col >= 7 and not has_langs:
                continue
            editable = is_fixed and col == 3
            calc = not editable
            apply_cell_style(
                ws.cell(row=row, column=col),
                fill=EDIT_FILL if editable else CALC_FILL,
                font=DATA_FONT,
                alignment=center if col != 4 else Alignment(wrap_text=True, vertical="top"),
                border=DATA_BORDER,
                locked=not editable,
            )
    set_col_widths(ws, (22, 8, 10, 44, 12, 12, 12, 12, 12))
    ws.freeze_panes = "A4"
    enable_sheet_protection(ws)


def build_questions_sheet(ws: Worksheet) -> None:
    ws.title = "Questions"
    ws.sheet_properties.tabColor = "C55A11"
    style_merged_title(ws, "A1", f"A1:{HIDDEN_LINE_COL}1")
    ws["A1"] = "Quiz steps - screens and scoring"
    ws["A2"] = (
        f"Add a step: fill the next empty row (column Artwork). "
        f"Rows {QUESTIONS_FIRST_ROW}-{QUESTIONS_LAST_ROW} are ready (max {MAX_QUESTION_SLOTS} steps). "
        "Question plays on = dropdown. Score grid: 0–10 per profile per button (B1–B4)."
    )
    style_instruction_row(ws, "A2", f"A2:{HIDDEN_LINE_COL}2")

    headers = {
        COL_QNUM: "Step",
        COL_ARTWORK: "Artwork",
        COL_QUESTION_ON: "Question plays on",
        COL_ARTWORK_ON: "Artwork plays on",
        COL_TEXT_PLAYER: "Text player",
        COL_ART_PLAYER: "Art player",
        COL_TEXT_ASPECT: "Text aspect",
        COL_ART_ASPECT: "Art aspect",
        COL_TEXT_FILE_FR: "Text file (FR)",
        COL_ART_FILE: "Art file",
    }
    for choice in range(NUM_CHOICES):
        for cat in range(NUM_CATEGORIES):
            col = COL_WEIGHTS + choice * NUM_CATEGORIES + cat
            headers[col] = f"B{choice + 1} {CAT_NAMES[cat]}"
    for col, title in headers.items():
        editable = col in (COL_ARTWORK, COL_QUESTION_ON) or (
            COL_WEIGHTS <= col <= COL_WEIGHTS_LAST
        )
        apply_cell_style(
            ws.cell(row=3, column=col, value=title),
            fill=questions_sheet_fill(col, editable=editable, calc=False, header=True),
            font=HEADER_FONT,
            locked=True,
        )

    defaults = list(DEFAULT_QUESTIONS)
    for slot in range(MAX_QUESTION_SLOTS):
        row = QUESTIONS_FIRST_ROW + slot
        if slot < len(defaults):
            default_cats, question_on, artwork = defaults[slot]
            fill_question_row(
                ws, row,
                default_cats=default_cats,
                question_on=question_on,
                artwork=artwork,
            )
        else:
            fill_question_row(ws, row, question_on="Portrait")

    ws.column_dimensions[get_column_letter(COL_TEXT_OFFSET)].hidden = True
    ws.column_dimensions[get_column_letter(COL_IMAGE_INDEX)].hidden = True
    ws.column_dimensions[HIDDEN_LINE_COL].hidden = True

    q_col = get_column_letter(COL_QUESTION_ON)
    screen_range = f"{q_col}{QUESTIONS_FIRST_ROW}:{q_col}{QUESTIONS_LAST_ROW}"
    add_list_validation(ws, screen_range, SCREEN_LIST)
    weight_col_start = get_column_letter(COL_WEIGHTS)
    weight_col_end = get_column_letter(COL_WEIGHTS_LAST)
    dv_weight = DataValidation(
        type="whole",
        operator="between",
        formula1=str(SCORE_WEIGHT_MIN),
        formula2=str(SCORE_WEIGHT_MAX),
        allow_blank=True,
    )
    dv_weight.error = f"Enter an integer from {SCORE_WEIGHT_MIN} to {SCORE_WEIGHT_MAX}."
    dv_weight.add(
        f"{weight_col_start}{QUESTIONS_FIRST_ROW}:{weight_col_end}{QUESTIONS_LAST_ROW}"
    )
    ws.add_data_validation(dv_weight)

    center = Alignment(horizontal="center", vertical="center")
    for row in range(QUESTIONS_FIRST_ROW, QUESTIONS_LAST_ROW + 1):
        for col in range(1, COL_WEIGHTS_LAST + 1):
            editable = col in (COL_ARTWORK, COL_QUESTION_ON) or (
                COL_WEIGHTS <= col <= COL_WEIGHTS_LAST
            )
            calc = col in (
                COL_QNUM, COL_TEXT_OFFSET, COL_ARTWORK_ON, COL_IMAGE_INDEX,
                COL_TEXT_PLAYER, COL_ART_PLAYER, COL_TEXT_ASPECT, COL_ART_ASPECT,
                COL_TEXT_FILE_FR, COL_ART_FILE,
            )
            apply_cell_style(
                ws.cell(row=row, column=col),
                fill=questions_sheet_fill(col, editable=editable, calc=calc, header=False),
                font=DATA_FONT,
                alignment=center if col != COL_ARTWORK else Alignment(horizontal="left"),
                border=DATA_BORDER,
                locked=not editable,
            )

    # A–L: layout + file names. M–AB: score grid.
    widths_before_weights = [6, 34, 18, 18, 8, 8, 12, 12, 14, 14, 16, 16]
    widths_score_grid = [16] * (COL_WEIGHTS_LAST - COL_WEIGHTS + 1)
    set_col_widths(ws, widths_before_weights + widths_score_grid)
    ws.freeze_panes = "A4"
    enable_sheet_protection(ws)


def build_codegen_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("CodeGen")
    ws.sheet_properties.tabColor = "BF8F00"
    style_merged_title(ws, "A1", "A1:E1")
    ws["A1"] = "QuestionnaireConfig.h generator"
    ws["A2"] = (
        "Copy column A from BEGIN through END into QuestionnaireConfig.h. "
        "Lines pull live from Questions (scores, screen, step). Skip blank array lines. "
        "NUM_QUESTIONS = COUNTA(Artwork). Recalculate the sheet after edits (F9 in Excel)."
    )
    style_instruction_row(ws, "A2", "A2:E2")
    ws.cell(row=CODEGEN_BEGIN_ROW, column=1, value="// --- BEGIN_QUESTIONNAIRE_CONFIG ---")
    ws.cell(row=CODEGEN_NUM_QUESTIONS_ROW, column=1, value=num_questions_cpp_formula())
    ws.cell(row=CODEGEN_ARRAY_OPEN_ROW, column=1, value='="static const QuestionEntry QUESTIONS[NUM_QUESTIONS] = {"')
    for slot in range(MAX_QUESTION_SLOTS):
        ws.cell(
            row=CODEGEN_FIRST_Q_ROW + slot,
            column=1,
            value=codegen_question_ref(QUESTIONS_FIRST_ROW + slot),
        )
    ws.cell(row=CODEGEN_CLOSE_ROW, column=1, value='="};"')
    ws.cell(row=CODEGEN_END_MARKER_ROW, column=1, value="// --- END_QUESTIONNAIRE_CONFIG ---")
    ws.cell(row=CODEGEN_END_MARKER_ROW + 1, column=1, value=(
        "NUM_QUESTIONS = count of rows with Artwork on Questions sheet. "
        "SD indices: text and artwork share the same step number on each card (0 = idle). "
        "Text filenames use language blocks on the text player only."
    ))
    for row in range(COPY_BLOCK_FIRST_ROW, CODEGEN_END_MARKER_ROW + 2):
        apply_cell_style(
            ws.cell(row=row, column=1),
            fill=COPY_FILL,
            font=Font(name="Consolas", size=10),
            alignment=Alignment(wrap_text=True, vertical="top"),
            locked=True,
        )
    set_col_widths(ws, (110,))
    ws.freeze_panes = "A5"
    enable_sheet_protection(ws)


def build_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    build_guide_sheet(wb.create_sheet("Guide"))
    build_questions_sheet(wb.create_sheet("Questions"))
    build_wiz_sd_sheet(
        wb.create_sheet("SD WIZ1"),
        title="SD WIZ1",
        wiz_name="WIZ1 portrait display",
        wiz_player="WIZ1",
        fixed_rows=WIZ1_FIXED_ROWS,
        tab_color="548235",
        first_row=WIZ1_FIRST_ROW,
        last_row=WIZ1_LAST_ROW,
    )
    build_wiz_sd_sheet(
        wb.create_sheet("SD WIZ2"),
        title="SD WIZ2",
        wiz_name="WIZ2 landscape display",
        wiz_player="WIZ2",
        fixed_rows=WIZ2_FIXED_ROWS,
        tab_color="7030A0",
        first_row=WIZ2_FIRST_ROW,
        last_row=WIZ2_LAST_ROW,
    )
    build_codegen_sheet(wb)
    return wb


def screen_to_cpp(screen: str) -> str:
    s = (screen or "").strip()
    return "Landscape" if s.lower().startswith("land") else "Portrait"


def iter_question_rows_from_sheet(ws: Worksheet) -> List[int]:
    """Row numbers on Questions sheet that have Artwork filled in."""
    rows: List[int] = []
    for row in range(QUESTIONS_FIRST_ROW, QUESTIONS_LAST_ROW + 1):
        artwork = ws.cell(row=row, column=COL_ARTWORK).value
        if artwork is not None and str(artwork).strip():
            rows.append(row)
    return rows


def read_choice_weights(ws: Worksheet, row: int) -> List[Tuple[int, ...]]:
    """Read 4×4 weight grid; fall back to DEFAULT_QUESTIONS template for blank cells."""
    slot = row - QUESTIONS_FIRST_ROW
    default_cats: Sequence[int] | None = None
    if 0 <= slot < len(DEFAULT_QUESTIONS):
        default_cats = DEFAULT_QUESTIONS[slot][0]
    rows: List[Tuple[int, ...]] = []
    for choice in range(NUM_CHOICES):
        cells: List[int | None] = []
        for cat in range(NUM_CATEGORIES):
            val = ws.cell(row=row, column=COL_WEIGHTS + choice * NUM_CATEGORIES + cat).value
            try:
                cells.append(int(val))
            except (TypeError, ValueError):
                cells.append(None)
        if all(v is None for v in cells):
            if default_cats is not None:
                rows.append(default_weights_for_primary(default_cats[choice]))
            else:
                rows.append(zero_weights_for_choice())
            continue
        filled: List[int] = []
        for i, v in enumerate(cells):
            if v is None:
                if default_cats is not None:
                    w = default_weights_for_primary(default_cats[choice])
                    filled.append(w[i])
                else:
                    filled.append(0)
            else:
                filled.append(max(SCORE_WEIGHT_MIN, min(SCORE_WEIGHT_MAX, v)))
        rows.append(tuple(filled))
    return rows


def emit_cpp_from_workbook(path: Path) -> str:
    wb = load_workbook(path, data_only=True)
    ws = wb["Questions"]
    question_rows = iter_question_rows_from_sheet(ws)
    n = len(question_rows)
    lines = [
        "// --- BEGIN_QUESTIONNAIRE_CONFIG ---",
        f"static const uint8_t NUM_QUESTIONS = {n};",
        "static const QuestionEntry QUESTIONS[NUM_QUESTIONS] = {",
    ]
    for step, row in enumerate(question_rows, start=1):
        screen = screen_to_cpp(str(ws.cell(row=row, column=COL_QUESTION_ON).value))
        artwork = ws.cell(row=row, column=COL_ARTWORK).value or f"Q{step}"
        off = ws.cell(row=row, column=COL_TEXT_OFFSET).value
        try:
            off = int(off)
        except (TypeError, ValueError):
            off = ws.cell(row=row, column=COL_QNUM).value
            try:
                off = int(off)
            except (TypeError, ValueError):
                off = step
        weights = read_choice_weights(ws, row)
        lines.append(f"  // Step {step} - {artwork}")
        lines.append(cpp_question_line(off, screen, weights))
    lines += ["};", "// --- END_QUESTIONNAIRE_CONFIG ---"]
    return "\n".join(lines)


def main(argv: Sequence[str] | None = None) -> None:
    repo = Path(__file__).resolve().parents[1]
    default_out = repo / "questions" / "MediaMap_workbook.xlsx"
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("-o", "--output", type=Path, default=default_out)
    parser.add_argument("--emit-cpp", action="store_true")
    args = parser.parse_args(list(argv) if argv is not None else None)

    if args.emit_cpp:
        if not args.output.exists():
            raise SystemExit(f"workbook not found: {args.output}")
        print(emit_cpp_from_workbook(args.output))
        return

    wb = build_workbook()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
