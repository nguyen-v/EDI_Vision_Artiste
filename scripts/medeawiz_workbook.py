#!/usr/bin/env python3
"""Resolve MedeaWiz clip targets from MediaMap_workbook.xlsx (Questions sheet)."""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

_SCRIPT_DIR = Path(__file__).resolve().parent
_REPO = _SCRIPT_DIR.parent
_DEFAULT_WORKBOOK = _REPO / "questions" / "MediaMap_workbook.xlsx"

# Import shared constants from generate_media_workbook.py
import importlib.util

_spec = importlib.util.spec_from_file_location(
    "media_workbook", _SCRIPT_DIR / "generate_media_workbook.py"
)
_mw = importlib.util.module_from_spec(_spec)
assert _spec.loader is not None
_spec.loader.exec_module(_mw)

FILES_PER_LANGUAGE = _mw.FILES_PER_LANGUAGE
LANG_CODES: tuple[str, ...] = _mw.LANG_CODES
PORTRAIT_ASPECT = _mw.PORTRAIT_ASPECT
LANDSCAPE_ASPECT = _mw.LANDSCAPE_ASPECT
COL_QNUM = _mw.COL_QNUM
COL_ARTWORK = _mw.COL_ARTWORK
COL_QUESTION_ON = _mw.COL_QUESTION_ON
COL_TEXT_OFFSET = _mw.COL_TEXT_OFFSET
COL_IMAGE_INDEX = _mw.COL_IMAGE_INDEX
QUESTIONS_FIRST_ROW = _mw.QUESTIONS_FIRST_ROW
placement = _mw.placement
read_question_entries_from_ws = _mw.read_question_entries_from_ws

Part = Literal["text", "image"]
ClipKind = Literal["idle", "question", "profile"]

PROFILE_ALIASES = {
    "emotions": 17,
    "emotion": 17,
    "emo": 17,
    "realiste": 18,
    "realist": 18,
    "rea": 18,
    "matiere": 19,
    "matter": 19,
    "mat": 19,
    "conteur": 20,
    "storyteller": 20,
    "con": 20,
}


@dataclass(frozen=True)
class ClipTarget:
    filename: str
    file_index: int
    sd_folder: str  # sd_wiz1 | sd_wiz2
    player: str  # WIZ1 | WIZ2
    encode_mode: str  # portrait | landscape (ffmpeg script mode)
    clip_type: ClipKind
    language: str | None
    step: int | None
    artwork: str | None
    question_on: str | None

    @property
    def relative_path(self) -> str:
        return f"{self.sd_folder}/{self.filename}"


def _lang_index(lang: str) -> int:
    key = lang.strip().upper()
    if key not in LANG_CODES:
        raise ValueError(f"unknown language {lang!r} (use {', '.join(LANG_CODES)})")
    return LANG_CODES.index(key)


def _encode_mode_for_aspect(aspect: str) -> str:
    return "portrait" if aspect == PORTRAIT_ASPECT else "landscape"


def _filename(file_index: int, language: str | None) -> str:
    idx = file_index
    if language is not None:
        idx = _lang_index(language) * FILES_PER_LANGUAGE + file_index
    return f"{idx:03d}.mp4"


def _find_question(
    entries: list[tuple[int, str, str, int, int]], step: int
) -> tuple[int, str, str, int, int]:
    for entry in entries:
        if entry[0] == step:
            return entry
    raise ValueError(f"question step {step} not found in workbook (no Artwork row?)")


def resolve_idle(part: Part, language: str | None) -> ClipTarget:
    if part == "text":
        if language is None:
            raise ValueError("idle text requires --lang (EN, FR, DE, IT)")
        return ClipTarget(
            filename=_filename(0, language),
            file_index=0,
            sd_folder="sd_wiz1",
            player="WIZ1",
            encode_mode="portrait",
            clip_type="idle",
            language=language.upper(),
            step=None,
            artwork="Idle intro text",
            question_on=None,
        )
    return ClipTarget(
        filename=_filename(0, None),
        file_index=0,
        sd_folder="sd_wiz2",
        player="WIZ2",
        encode_mode="landscape",
        clip_type="idle",
        language=None,
        step=None,
        artwork="Idle artwork",
        question_on=None,
    )


def resolve_profile(part: Part, profile: str, language: str | None) -> ClipTarget:
    key = profile.strip().lower()
    if key not in PROFILE_ALIASES:
        names = ", ".join(sorted({k for k in PROFILE_ALIASES if len(k) > 3}))
        raise ValueError(f"unknown profile {profile!r} (use {names})")
    file_index = PROFILE_ALIASES[key]
    if part == "text":
        if language is None:
            raise ValueError("profile text requires --lang (EN, FR, DE, IT)")
        return ClipTarget(
            filename=_filename(file_index, language),
            file_index=file_index,
            sd_folder="sd_wiz1",
            player="WIZ1",
            encode_mode="portrait",
            clip_type="profile",
            language=language.upper(),
            step=None,
            artwork=f"Profile {profile}",
            question_on=None,
        )
    return ClipTarget(
        filename=_filename(file_index, None),
        file_index=file_index,
        sd_folder="sd_wiz2",
        player="WIZ2",
        encode_mode="landscape",
        clip_type="profile",
        language=None,
        step=None,
        artwork=f"Profile art {profile}",
        question_on=None,
    )


def resolve_question(
    entries: list[tuple[int, str, str, int, int]],
    step: int,
    part: Part,
    language: str | None,
) -> ClipTarget:
    _step, question_on, artwork, text_offset, image_index = _find_question(entries, step)
    text_wiz, art_wiz, text_asp, art_asp = placement(question_on)

    if part == "text":
        if language is None:
            raise ValueError(f"Q{step} text requires --lang (EN, FR, DE, IT)")
        player = text_wiz
        aspect = text_asp
        file_index = text_offset
    else:
        if language is not None:
            raise ValueError("question artwork is language-independent (omit --lang)")
        player = art_wiz
        aspect = art_asp
        file_index = image_index

    sd_folder = "sd_wiz1" if player == "WIZ1" else "sd_wiz2"
    return ClipTarget(
        filename=_filename(file_index, language if part == "text" else None),
        file_index=file_index,
        sd_folder=sd_folder,
        player=player,
        encode_mode=_encode_mode_for_aspect(aspect),
        clip_type="question",
        language=language.upper() if language else None,
        step=step,
        artwork=artwork,
        question_on=question_on,
    )


def resolve_clip(
    *,
    workbook: Path,
    question: int | None = None,
    idle: bool = False,
    profile: str | None = None,
    part: Part,
    language: str | None = None,
) -> ClipTarget:
    n_slots = (
        (1 if question is not None else 0)
        + (1 if idle else 0)
        + (1 if profile is not None else 0)
    )
    if n_slots != 1:
        raise ValueError("specify exactly one of --q STEP, --idle, or --profile NAME")

    if idle:
        return resolve_idle(part, language)
    if profile is not None:
        return resolve_profile(part, profile, language)

    assert question is not None
    wb = load_workbook(workbook, data_only=True)
    if "Questions" not in wb.sheetnames:
        raise ValueError(f"Questions sheet missing in {workbook}")
    entries = read_question_entries_from_ws(wb["Questions"])
    if not entries:
        raise ValueError(f"no filled question rows in {workbook}")
    return resolve_question(entries, question, part, language)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=_DEFAULT_WORKBOOK,
        help=f"MediaMap workbook (default: {_DEFAULT_WORKBOOK.relative_to(_REPO)})",
    )
    slot = parser.add_mutually_exclusive_group(required=True)
    slot.add_argument("--q", "--question", type=int, dest="question", metavar="STEP")
    slot.add_argument("--idle", action="store_true")
    slot.add_argument("--profile", metavar="NAME", help="emotions, realiste, matiere, conteur")
    parser.add_argument(
        "--part",
        required=True,
        choices=("text", "image"),
        help="text (per language) or image/artwork",
    )
    parser.add_argument(
        "--lang",
        "--language",
        dest="language",
        help="EN, FR, DE, IT (required for text clips)",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        help="If set with --shell, build OUTPUT under this directory",
    )
    parser.add_argument("--json", action="store_true", help="Print JSON (default)")
    parser.add_argument(
        "--shell",
        action="store_true",
        help="Print shell assignments MODE, TYPE, OUTPUT for convert_image_to_medeawiz.sh",
    )
    args = parser.parse_args(argv)

    try:
        target = resolve_clip(
            workbook=args.workbook,
            question=args.question,
            idle=args.idle,
            profile=args.profile,
            part=args.part,
            language=args.language,
        )
    except (ValueError, OSError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1

    if args.shell:
        out_path = target.relative_path
        if args.out_dir is not None:
            out_path = str(args.out_dir / target.sd_folder / target.filename)
        # shell-safe single quotes
        def sh(s: str) -> str:
            return "'" + s.replace("'", "'\"'\"'") + "'"

        print(f"MODE={sh(target.encode_mode)}")
        print(f"TYPE={sh(target.clip_type)}")
        print(f"OUTPUT={sh(out_path)}")
        print(f"SD_FOLDER={sh(target.sd_folder)}")
        print(f"FILENAME={sh(target.filename)}")
        print(f"PLAYER={sh(target.player)}")
        if target.language:
            print(f"LANGUAGE={sh(target.language)}")
        if target.step is not None:
            print(f"STEP={target.step}")
        if target.artwork:
            print(f"ARTWORK={sh(target.artwork)}")
        return 0

    print(json.dumps(asdict(target), indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
